import logging
import urllib3.contrib.appengine
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters
from openpyxl import load_workbook
import re
from telegram import ReplyKeyboardMarkup, KeyboardButton

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)

api_token = None

cities = ['Astana', 'Almaty', 'Shymkent']
positions = ['Intern', 'Senior', 'Junior']

wb = load_workbook('database.xlsx')
ws = wb['Сотрудник']

def start(update, context):
    update.message.reply_text("Hello! Please enter your IIN code.")
    context.user_data['state'] = 'waiting_iin'

def handle_input(update, context):
    state = context.user_data.get('state')
    message = update.message.text
    if state == 'waiting_iin':
        if not re.match(r'^\d{12}$', message):
            update.message.reply_text("Invalid IIN code. Please enter a 12-digit number.")
            return
        if check_duplicate_iin(message):
            update.message.reply_text("This IIN code already exists in the database. Please enter a different one.")
            return
        context.user_data['iin_code'] = update.message.text
        update.message.reply_text("Please enter your surname.")
        context.user_data['state'] = 'waiting_surname'
    elif state == 'waiting_surname':
        context.user_data['surname'] = update.message.text
        update.message.reply_text("Please enter your name.")
        context.user_data['state'] = 'waiting_name'
    elif state == 'waiting_name':
        context.user_data['name'] = update.message.text
        update.message.reply_text("Please enter your father's name.")
        context.user_data['state'] = 'waiting_father_name'
    elif state == 'waiting_father_name':
        context.user_data['father_name'] = update.message.text
        keyboard = [[KeyboardButton(city)] for city in cities]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True)
        update.message.reply_text("Please choose your city:", reply_markup=reply_markup)
        context.user_data['state'] = 'waiting_city'
    elif state == 'waiting_city':
        context.user_data['city'] = update.message.text
        update.message.reply_text('Please enter your phone number.')
        context.user_data['state'] = 'waiting_phone_number'
    elif state == 'waiting_phone_number':
        message = update.message.text 
        if not re.match(r'^\+\d{11}$', message):
            update.message.reply_text("Invalid phone number. Please try again.")
            return
        context.user_data['phone_number'] = update.message.text
        keyboard = [[KeyboardButton(position)] for position in positions]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True)
        update.message.reply_text("Please select your position:", reply_markup=reply_markup)
        context.user_data['state'] = 'waiting_position'
    elif state == 'waiting_position':
        context.user_data['position'] = update.message.text
        update.message.reply_text("Please enter your id.")
        context.user_data['state'] = 'waiting_id'
    elif state == 'waiting_id':
        if not re.match(r'^\d{9}$', message):
            update.message.reply_text("Invalid ID. Please enter a 9-digit number.")
            return
        context.user_data['id'] = update.message.text
        update.message.reply_text("Please enter your IBAN.")
        context.user_data['state'] = 'waiting_iban'
    elif state == 'waiting_iban':
        if not re.match(r'^KZ\d{18}$', message):
            update.message.reply_text("Invalid IBAN. Please enter an IBAN according to Kazakhstan's format.")
            return
        context.user_data['iban'] = update.message.text
        save_to_excel(update, context)
        update.message.reply_text('Registration completed successfully.')
    
def check_duplicate_iin(iin_code):
    for row in ws.iter_rows(values_only=True):
        if row[0] == iin_code:
            return True
    return False

def save_to_excel(update, context):
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=context.user_data.get('iin_code'))
    ws.cell(row=row, column=2, value=context.user_data.get('surname'))
    ws.cell(row=row, column=3, value=context.user_data.get('name'))
    ws.cell(row=row, column=4, value=context.user_data.get('father_name'))
    ws.cell(row=row, column=5, value=context.user_data.get('city'))
    ws.cell(row=row, column=6, value=context.user_data.get('phone_number'))
    ws.cell(row=row, column=7, value=context.user_data.get('position'))
    ws.cell(row=row, column=8, value=context.user_data.get('id'))
    ws.cell(row=row, column=9, value=context.user_data.get('iban'))
    wb.save('database.xlsx')

# def handle_city_selection(update, context):
#     update.message.reply_text("Please enter your phone number.")
#     context.user_data['state'] = 'waiting_phone_number'

# def handle_position_selection(update, context):
#     update.message.reply_text("Please enter your ID.")
#     context.user_data['state'] = 'waiting_id'

def main():
    updater = Updater(api_token, use_context=True)
    dp = updater.dispatcher

    dp.add_handler(CommandHandler('start', start))
    dp.add_handler(MessageHandler(Filters.text & ~Filters.command, handle_input))

    updater.start_polling()
    updater.idle()

if __name__ == '__main__':
    main()

